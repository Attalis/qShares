import csv
import datetime
import os
import pickle

import pandas as pd
from openpyxl import Workbook

import settings
from data_access.tools.data import data_factory
from portfolio_tools.trading import Trade


class Portfolio(object):
    def __init__(self, logger=None):
        # list of trades
        self.trades = []
        self.pnl = []
        # self.perf_cols = ['fs_perm_sec_id', 'date', 'trade_date', 'stock_id', 'ticker_exchange', 'ex_date', 'dps_cents',
        #                           'franking_pct_series', 'close_price',
        #                           'shares', 'days', 'status', 'direction']
        # self.performance = pd.DataFrame(columns=self.perf_cols)
        # self.performance = self.performance.set_index(['fs_perm_sec_id', 'date'])

        # self.daily_pnl = []
        # self.cash_account = []

        self._data_path = settings.DATA_PATH
        self._output_path = settings.OUTPUT_PATH

        self._company_tax_rate = settings.COMPANY_TAX_RATE

        self._backtest_start_date = settings.BACKTEST_START_DATE

        # filter levels
        self._limit_dy = settings.LIMIT_DY
        self._limit_er = settings.LIMIT_ER
        self._limit_volatility = settings.LIMIT_VOLATILITY
        self._days_to_hold = settings.DAYS_TO_HOLD

        # Fund size values
        self._fund_seed_value = settings.FUND_SIZE
        self._bucket_weight_1 = settings.BUCKET_WEIGHT_1
        self._bucket_weight_2 = settings.BUCKET_WEIGHT_2
        self._bucket_weight_3 = settings.BUCKET_WEIGHT_3

        # Trade cost constants
        self._gross_exposure = settings.GROSS_EXPOSURE
        self._net_exposure = settings.NET_EXPOSURE
        self._BBSW = settings.BBSW
        self._trade_commission = settings.TRADE_COMMISSION
        self._long_funding_cost_BBSW = settings.LONG_FUNDING_COST_BBSW
        self._short_cost = settings.SHORT_COST

        self._data_factory = data_factory()

        self._dataset = self._data_factory.build_hist_data()
        self._dataset = self._data_factory.calc_factors()
        self._prices = self._data_factory.get_prices_from_csv()
        self.prices = self._data_factory.build_prices(use_cache=True)

        self._fund_seed_value = settings.FUND_SIZE
        self._cash_account = self._fund_seed_value
        self._franking_account = 0
        self.event_count = 0

        self.max_open_trades = 0

        self.perf = pd.DataFrame()

        # add the start position

    # properties
    def trade_dates(self):
        # Get the dates to process for the portfolio (dates are in a list)
        tdates = pd.to_datetime(self._dataset.index_date.unique()).sort_values()
        tdates = tdates.strftime('%Y-%m-%d').tolist()
        return tdates

    # Portfolio Performance statistics
    def value(self):
        total = 0
        for t in self.trades():
            total += t.close_price * t.shares

        return total

    def max_stocks_held(self):
        pass

    def avg_dollar_return_per_trade(self):
        pass

    def avg_pct_return_per_trade(self):
        pass

    def annualised_return(self):
        pass

    def information_ratio(self):
        pass

    def annualised_downside_volatility(self):
        pass

    def max_draw_down(self):
        pass

    def stock_hit_rate(self):
        pass

    def avg_net_exposure(self):
        pass

    def max_net_exposure(self):
        pass

    def performance_summary(self):

        # calculate portfolio performance
        print('####################################################################')
        print('###################  Summary                  ######################')
        print('####################################################################')
        print('Start: ' + '${:,.0f}'.format(self._fund_seed_value))
        print('End  : ' + '${:,.0f}'.format(self._portfolio_return))
        print('Franking  : ' + '${:,.0f}'.format(self._franking_account))
        print('####################################################################')
        print('####################################################################')
        print('####################################################################')


        # [all_trades.ex_date.tolist(), all_trades.groupby(all_trades['ex_date'].map(lambda x: x.day)).count()]
        # y = all_trades.groupby(all_trades['ex_date'].map(lambda x: x.day)).count()['stock_id'].tolist()

    # methods
    def save(self, filename=''):
        if filename == '':
            filename = os.path.join(settings.OUTPUT_PATH, settings.DEFAULT_PORT_FILE.split('.')[0]
                                    + datetime.datetime.today().strftime('%Y-%m-%d') + '.p')
        with open(filename, 'wb') as pfile:
            pickle.dump(self.trades, pfile)

    def load(self, filename=''):
        if filename == '':
            filename = os.path.join(settings.OUTPUT_PATH, settings.DEFAULT_PORT_FILE.split('.')[0]
                                    + datetime.datetime.today().strftime('%Y-%m-%d') + '.p')
        self.trades = pickle.load(open(filename, 'rb'))

    def apply_strategy_filters(self, sample_data):
        sample_data = sample_data.query('(has_div == 1)')
        sample_data = sample_data.query('(er > ' + str(settings.LIMIT_ER) + ')')
        # sample_data = sample_data.query('(hist_dy_series > ' + str(settings.LIMIT_DY) + ')')
        # sample_data = sample_data.query('(ann_dy > ' + str(settings.LIMIT_DY) + ')')
        # TODO: Do we need to filter on DY
        # sample_data = sample_data.query('(ann_dy > 0.01)')

        return sample_data

    def get_trade_size_adjustment(self, franking_pct, div_yield):
        # TODO check that annualised DY is used and adjust to 5% threshold

        if (franking_pct == 1) and (div_yield >= self._limit_dy):
            return 0.075

        if (franking_pct == 1) and (div_yield <= self._limit_dy):
            return 0.05

        # TODO: check this logic with JS
        if (franking_pct <= 1) and (franking_pct > 0) and (div_yield >= self._limit_dy):
            return 0.015
        else:
            return 0.015

    def open_trades(self, sample_data):

        divs_data = sample_data.query('(dps_cents > 0)')

        if not divs_data.empty:
            # process each row separately
            for ix, tr in divs_data.iterrows():
                self.event_count += 1
                # div found, test for trade
                frank_pct = tr['franking_pct_series']
                div_yield = tr['ann_dy']
                # trade_size_adjustment = self.get_trade_size_adjustment(tr['franking_pct_series'], tr['hist_dy_series'])
                trade_size_adjustment = self.get_trade_size_adjustment(frank_pct, div_yield)
                if trade_size_adjustment > 0:
                    trade_value = self._fund_seed_value * trade_size_adjustment
                    # trade_value = self._cash_account * trade_size_adjustment
                    # TODO: use the cash balance from previous 1/12 or 1/6 to provide the size of the trade
                    # fund size must increase over time
                    # take the fund size value from the total fund cash balance (ex franking)
                    # from the previous 1/12 or 1/6
                    # this should keep the total number of trades fairly even at 60-80 per season


                    # Check to see if there's enough cash balance to do the trade
                    if self._cash_account - trade_value > 0:
                        print('Status: ' + '          : ' + 'open trade  : ' + '${:,.0f}'.format(self._cash_account)
                              + ' : trade size : ' + '${:,.0f}'.format(trade_value)
                              + ': ' + tr[['ticker_exchange']][0])
                        self._cash_account -= trade_value

                        stock_id = tr['stock_id']
                        exchange_ticker = tr['ticker_exchange']
                        div_ex_date = tr['ex_date']
                        trade_open_date = (div_ex_date + datetime.timedelta(days=-(self._days_to_hold))).to_pydatetime()
                        close_price = self._data_factory.get_price(trade_open_date.strftime('%Y-%m-%d'), stock_id)
                        div_dollars = tr['dps_cents'] / 100

                        # Build the main long trade and add to collection
                        divtrade = Trade(stock_id, exchange_ticker, trade_open_date, close_price)
                        divtrade.direction = 'long'
                        divtrade.div_ex_date = div_ex_date
                        divtrade.div_pershare = div_dollars
                        divtrade.franking_pct = frank_pct
                        divtrade.dy = div_yield
                        divtrade.trade_size = trade_value
                        divtrade.trade_size_adj = trade_size_adjustment
                        self.trades.append(divtrade)
                        fs_perm_sec_id = stock_id

                        # row = pd.DataFrame([[fs_perm_sec_id, trade_open_date, trade_open_date, stock_id, exchange_ticker, div_ex_date, div_dollars, frank_pct, close_price, divtrade.shares(),  divtrade.days_open, divtrade.status, divtrade.direction]], columns=self.perf_cols)
                        # row = row.set_index(['fs_perm_sec_id', 'date'])
                        # self.performance = self.performance.append(row)

                        if frank_pct == 1:
                            # if there's a franking build the long/short trades
                            long_trade_size = (trade_value / 0.3 - trade_value) * 1
                            ls_long_trade = Trade(stock_id, exchange_ticker, trade_open_date, close_price)
                            ls_long_trade.direction = 'long'
                            ls_long_trade.div_ex_date = div_ex_date
                            ls_long_trade.div_pershare = div_dollars
                            ls_long_trade.franking_pct = frank_pct
                            ls_long_trade.dy = div_yield
                            ls_long_trade.trade_size = long_trade_size
                            ls_long_trade.trade_size_adj = trade_size_adjustment
                            self.trades.append(ls_long_trade)
                            # row = pd.DataFrame([[fs_perm_sec_id, trade_open_date, trade_open_date, stock_id, exchange_ticker, div_ex_date, div_dollars, frank_pct, close_price, ls_long_trade.shares(), ls_long_trade.days_open, ls_long_trade.status, ls_long_trade.direction]], columns=self.perf_cols)
                            # row = row.set_index(['fs_perm_sec_id', 'date'])
                            # self.performance = self.performance.append(row)

                            # and the short side
                            short_trade_size = long_trade_size
                            ls_short_trade = Trade(stock_id, exchange_ticker, trade_open_date, close_price)
                            ls_short_trade.direction = 'short'
                            ls_short_trade.div_ex_date = div_ex_date
                            ls_short_trade.div_pershare = div_dollars
                            ls_short_trade.franking_pct = frank_pct
                            ls_short_trade.dy = div_yield
                            ls_short_trade.trade_size = short_trade_size
                            ls_short_trade.trade_size_adj = trade_size_adjustment
                            self.trades.append(ls_short_trade)
                            # row = pd.DataFrame([[fs_perm_sec_id, trade_open_date, trade_open_date, stock_id, exchange_ticker, div_ex_date, div_dollars, frank_pct, close_price, ls_short_trade.shares(), ls_short_trade.days_open, ls_short_trade.status, ls_short_trade.direction]], columns=self.perf_cols)
                            # row = row.set_index(['fs_perm_sec_id', 'date'])
                            # self.performance = self.performance.append(row)

    def update_trades(self, trade_date):
        tdate = datetime.datetime.strptime(trade_date, '%Y-%m-%d')
        for t in self.trades:
            # if t.days_open <= self._days_to_hold:
            if tdate >= (t.div_ex_date + datetime.timedelta(days=1)) and (t.status == 'open'):
                close_price = self._data_factory.get_price(tdate.strftime('%Y-%m-%d'), t.stock_id)
                t.close(tdate, close_price)

                # update the portfolio cash balance to put back the return for the stock plus the div
                self._cash_account += t.cash_at_end()
                self._franking_account += t.franking_credit()

                print('Status: ' + '          : ' + 'close trade : ' + '${:,.0f}'.format(self._cash_account)
                      + ' : trade size : ' + '${:,.0f}'.format(t.trade_size) + ': ' + t.exchange_ticker)

                # TODO reinvest trades less than max days
                # is it possible to include the funding costs each day
                # add in the div on the close of the ex date
                # if there's balance and theres an ex date more than 2 days away and stock meets entry criteria
                # and hasn't been traded yet this season then open trade and hold for 47 days

    def build_pnl(self):
        # take the trade list and build out the
        # walk through dates
        for trade_date in self.trade_dates():
            print('')
            print(trade_date, end="", flush = True)
            closed_trade_count = 0
            open_trade_count = 0
            tdate = datetime.datetime.strptime(trade_date, '%Y-%m-%d')
            # walk through trades and only look at the closed trades.
            for t in self.trades:
                if (t.status == 'closed'):
                    if t.open_date <= tdate <= t.close_date:
                        closed_trade_count += 1
                        print('.', end="", flush = True)
                        try:
                            close_price = self.prices[t.stock_id].loc[tdate].close_price[0]
                        except:
                            close_price = self.prices[t.stock_id].loc[tdate].close_price
                        # row_data = [trade_date, t.stock_id, t.exchange_ticker, str(t.shares()), str(t.start_price),
                        #             str(close_price)]

                        row_data = [trade_date, t.stock_id, t.exchange_ticker, t.open_date, str(t.start_price),
                                    t.close_date, t.end_price,
                                    str(t.trade_size),
                                    str(t.days_open), t.direction, t.status, t.div_ex_date, str(t.div_pershare),
                                    str(t.franking_pct),
                                    str(t.dy), str(t.trade_size_adj), str(t.shares()), str(t.shares() * t.start_price),
                                    str(t.franking_credit()), str(t.cash_dividend()),
                                    str(t.cash_dividend() / t.start_price),
                                    str(t.tran_cost()), str(t.fund_cost()), str(t.short_cost()),
                                    str(t.cash_dividend() * t.shares()), str(close_price)
                                    ]

                        self.pnl.append(row_data)
                elif t.status == 'open':
                    if t.open_date <= tdate:
                        open_trade_count += 1
                        print('*', end="", flush = True)
                        try:
                            close_price = self.prices[t.stock_id].loc[tdate].close_price[0]
                        except:
                            close_price = self.prices[t.stock_id].loc[tdate].close_price

                        # row_data = [trade_date, t.stock_id, t.exchange_ticker, str(t.shares()), str(t.start_price),
                        #             str(close_price)]

                        row_data = [trade_date, t.stock_id, t.exchange_ticker, t.open_date, str(t.start_price),
                                    t.close_date, t.end_price,
                                    str(t.trade_size),
                                    str(t.days_open), t.direction, t.status, t.div_ex_date, str(t.div_pershare),
                                    str(t.franking_pct),
                                    str(t.dy), str(t.trade_size_adj), str(t.shares()), str(t.shares() * t.start_price),
                                    str(t.franking_credit()), str(t.cash_dividend()),
                                    str(t.cash_dividend() / t.start_price),
                                    str(t.tran_cost()), str(t.fund_cost()), str(t.short_cost()),
                                    str(t.cash_dividend() * t.shares()), str(close_price)
                                    ]

                        self.pnl.append(row_data)
            if closed_trade_count > self.max_open_trades:
                self.max_open_trades = closed_trade_count

            print(': O: ' + str(open_trade_count) + ' C: ' + str(closed_trade_count), end="", flush = True)

            print('')

        # def build_perf(self):
        #
        #     for trade_date in self.trade_dates():
        #         self.performance.query('(date == "' + trade_date + '")')

        # def get_last_pnl(self):
        #     pnl_rows = []
        #     # get the pnl balance
        #     if len(self.pnl) > 0:
        #         for p in self.pnl:
        #             pnl_rows.append(p)
        #
        #     return pnl_rows
        #
        # def update_pnl(self, trade_date):
        #
        #     last_rows = self.get_last_pnl()
        #
        #     tdate = datetime.datetime.strptime(trade_date, '%Y-%m-%d')
        #
        #     if len(self.trades) > 0:
        #
        #         # walk through trades and only look at the closed trades.
        #         for t in self.trades:
        #             if (t.status == 'closed'):
        #                 if (t.open_date <= tdate <= t.close_date):
        #                     close_price = self._data_factory.get_close_price(tdate, t.stock_id)
        #                     row_data = [trade_date, t.stock_id, t.exchange_ticker, str(t.shares()), str(close_price),
        #                                 str(t.shares() * close_price)]
        #                     self.pnl.append(row_data)
        #     else:  # there's no trades so roll the balance forward
        #         # if there's nothing in the pnl then use the cash balance to create a record
        #         if len(self.pnl) == 0:
        #             row_data = [trade_date, self._fund_seed_value]
        #             self.pnl.append(row_data)
        #         else:
        #             pass

        def pnl_to_csv(self, filename=''):

            if filename == '':
                filename = os.path.join(settings.OUTPUT_PATH, settings.DEFAULT_PNL_FILE.split('.')[0]
                                        + datetime.datetime.today().strftime('%Y-%m-%d') + '.csv')
            with open(filename, 'w') as pnlfile:
                csvwriter = csv.writer(pnlfile, dialect='excel')
                # header = ['trade_date', 'stock_id', 'exchange_ticker', 'shares', 'trade_open_price', 'close_price']

                header = ['market_trade_date', 'stock_id', 'exchange_ticker', 'open_date', 'start_price', 'close_date',
                          'end_price',
                          'max_exposure', 'days_open', 'direction', 'status', 'div_ex_date', 'div_pershare',
                          'franking_pct', 'dy', 'trade_size_adj', 'shares',
                          'entry_value', 'franking_credit', 'gross_div',
                          'gross_yield', 'trading_costs', 'funding_costs', 'shorting_costs', 'cash_dividends',
                          # 'franking_credits', 'exit_value',
                          'market_close_price']
                csvwriter.writerow(header)
                # walk through dates

                for p in self.pnl:
                    csvwriter.writerow(p)

        def build_trades_list(self):
            # loop to process the historic data
            for trade_date in self.trade_dates():
                # Check age of existing trades and close out if necessary
                print('Status: ' + trade_date + ': cash balance: ' + '${:,.0f}'.format(self._cash_account))
                # self.update_trades(trade_date)
                self.update_trades(trade_date)

                # Apply the filters (has_div, ER, volatility) to limit the universe
                sample_data = self.apply_strategy_filters(self._dataset.query('(date == "' + trade_date + '")'))

                # Check for new divs to trade and open trades
                self.open_trades(sample_data)

                filename = os.path.join(settings.OUTPUT_PATH, 'per-'
                                        + datetime.datetime.today().strftime('%Y-%m-%d') + '.csv')
                # self.performance.to_csv(filename)
                # self.update_pnl(trade_date)

        def trades_to_csv(self, filename=''):

            if filename == '':
                filename = os.path.join(settings.OUTPUT_PATH, settings.DEFAULT_TRADE_FILE.split('.')[0]
                                        + datetime.datetime.today().strftime('%Y-%m-%d') + '.csv')

            with open(filename, 'w') as tradesfile:
                csvwriter = csv.writer(tradesfile, dialect='excel')
                header = ['stock_id', 'exchange_ticker', 'open_date', 'start_price', 'close_date', 'end_price',
                          'max_exposure', 'days_open', 'direction', 'status', 'div_ex_date', 'div_pershare',
                          'franking_pct', 'dy', 'trade_size_adj', 'shares',
                          'entry_value', 'franking_credit', 'gross_div',
                          'gross_yield', 'trading_costs', 'funding_costs', 'shorting_costs', 'cash_dividends']
                # , 'franking_credits', 'exit_value']
                csvwriter.writerow(header)
                # walk through trades
                for t in self.trades:
                    row_data = [t.stock_id, t.exchange_ticker, t.open_date, str(t.start_price), t.close_date,
                                t.end_price,
                                str(t.trade_size),
                                str(t.days_open), t.direction, t.status, t.div_ex_date, str(t.div_pershare),
                                str(t.franking_pct),
                                str(t.dy), str(t.trade_size_adj), str(t.shares()), str(t.shares() * t.start_price),
                                str(t.franking_credit()), str(t.cash_dividend()),
                                str(t.cash_dividend() / t.start_price),
                                str(t.tran_cost()), str(t.fund_cost()), str(t.short_cost()),
                                str(t.cash_dividend() * t.shares())
                                ]
                    csvwriter.writerow(row_data)

        def summary(self):
            # TODO: Build Portfolio Performance Summary
            # write out time series of cash balance and franking balance daily
            # total long exposure (long + long_ls)
            # total short exposure (short_ls)
            # net exposure (total_long + total_short) (assuming total short is always negative)
            # num unique stocks open each day
            # daily pnl time series
            # Cash + franking

            wb = Workbook()
            ws = wb.create_sheet()
            # ws.append(['blah', 'blah'])


            self.max_open_trades
            self.max_draw_down()

            print('')
            print('#######################################################################################')
            print('#######################################################################################')
            # cols = ['trade_date', 'stock_id', 'exchange_ticker', 'shares', 'trade_open_price', 'close_price']
            cols = ['market_trade_date', 'stock_id', 'exchange_ticker', 'open_date', 'start_price', 'close_date',
                    'end_price',
                    'max_exposure', 'days_open', 'direction', 'status', 'div_ex_date', 'div_pershare',
                    'franking_pct', 'dy', 'trade_size_adj', 'shares',
                    'entry_value', 'franking_credit', 'gross_div',
                    'gross_yield', 'trading_costs', 'funding_costs', 'shorting_costs', 'cash_dividends',
                    #  'franking_credits', 'exit_value',
                    'market_close_price']

            self.perf = pd.DataFrame(self.pnl, columns=cols)
            print('Unique stocks traded', str(len(self.perf.exchange_ticker.unique())))
            print('Stocks traded', self.perf.exchange_ticker.unique())
            b = self.perf
            b['date'] = pd.to_datetime(b['market_trade_date'], format='%Y-%m-%d')
            b = b.set_index(pd.DatetimeIndex(b['date']))
            # s = b.groupby(by=['trade_date','stock_id']).count() # this gives me a count for each stock each day

            b['stock_id'].resample("M").count()
            b['year'] = b.date.dt.year
            c = b.drop_duplicates(subset=['year', 'exchange_ticker'])

            u = c.groupby(['year'])['stock_id'].count()

            print(u)

            # calc the forward return
            b['value'] = pd.to_numeric(b.market_close_price) * pd.to_numeric(b.shares)
            # b['forward_month_ret'] = b.groupby(level=0)['tret'].shift(1)
            d = b
            d['value_prev'] = d.groupby(level=0)['value'].shift(1)
            d['return'] = (d.value / d.groupby(level=0)['value'].shift(1)) - 1

            for year in b['year'].unique():
                # print(year)
                stocks = b.query('(year == ' + str(year) + ')')['exchange_ticker'].unique()
                print(year, stocks)

            wb.save('/Users/connahcutbush/Desktop/perf_summary.xlsx')

    def build_trades_list(self):
        # loop to process the historic data
        for trade_date in self.trade_dates():
            # Check age of existing trades and close out if necessary
            print('Status: ' + trade_date + ': cash balance: ' + '${:,.0f}'.format(self._cash_account))
            # self.update_trades(trade_date)
            self.update_trades(trade_date)

            # Apply the filters (has_div, ER, volatility) to limit the universe
            sample_data = self.apply_strategy_filters(self._dataset.query('(date == "' + trade_date + '")'))

            # Check for new divs to trade and open trades
            self.open_trades(sample_data)

            filename = os.path.join(settings.OUTPUT_PATH, 'per-'
                                    + datetime.datetime.today().strftime('%Y-%m-%d') + '.csv')
            # self.performance.to_csv(filename)
            # self.update_pnl(trade_date)

    def trades_to_csv(self, filename=''):

        if filename == '':
            filename = os.path.join(settings.OUTPUT_PATH, settings.DEFAULT_TRADE_FILE.split('.')[0]
                                    + datetime.datetime.today().strftime('%Y-%m-%d') + '.csv')

        with open(filename, 'w') as tradesfile:
            csvwriter = csv.writer(tradesfile, dialect='excel')
            header = ['stock_id', 'exchange_ticker', 'open_date', 'start_price', 'close_date', 'end_price',
                      'max_exposure', 'days_open', 'direction', 'status', 'div_ex_date', 'div_pershare',
                      'franking_pct', 'dy', 'trade_size_adj', 'shares',
                      'entry_value', 'franking_credit', 'gross_div',
                      'gross_yield', 'trading_costs', 'funding_costs', 'shorting_costs', 'cash_dividends']
            # , 'franking_credits', 'exit_value']
            csvwriter.writerow(header)
            # walk through trades
            for t in self.trades:
                row_data = [t.stock_id, t.exchange_ticker, t.open_date, str(t.start_price), t.close_date, t.end_price,
                            str(t.trade_size),
                            str(t.days_open), t.direction, t.status, t.div_ex_date, str(t.div_pershare),
                            str(t.franking_pct),
                            str(t.dy), str(t.trade_size_adj), str(t.shares()), str(t.shares() * t.start_price),
                            str(t.franking_credit()), str(t.cash_dividend()), str(t.cash_dividend() / t.start_price),
                            str(t.tran_cost()), str(t.fund_cost()), str(t.short_cost()),
                            str(t.cash_dividend() * t.shares())
                            ]
                csvwriter.writerow(row_data)

    def summary(self):
        # TODO: Build Portfolio Performance Summary
        # write out time series of cash balance and franking balance daily
        # total long exposure (long + long_ls)
        # total short exposure (short_ls)
        # net exposure (total_long + total_short) (assuming total short is always negative)
        # num unique stocks open each day
        # daily pnl time series
        # Cash + franking

        wb = Workbook()
        ws = wb.create_sheet()
        # ws.append(['blah', 'blah'])


        self.max_open_trades
        self.max_draw_down()

        print('')
        print('#######################################################################################')
        print('#######################################################################################')
        # cols = ['trade_date', 'stock_id', 'exchange_ticker', 'shares', 'trade_open_price', 'close_price']
        cols = ['market_trade_date', 'stock_id', 'exchange_ticker', 'open_date', 'start_price', 'close_date',
                'end_price',
                'max_exposure', 'days_open', 'direction', 'status', 'div_ex_date', 'div_pershare',
                'franking_pct', 'dy', 'trade_size_adj', 'shares',
                'entry_value', 'franking_credit', 'gross_div',
                'gross_yield', 'trading_costs', 'funding_costs', 'shorting_costs', 'cash_dividends',
                #  'franking_credits', 'exit_value',
                'market_close_price']

        self.perf = pd.DataFrame(self.pnl, columns=cols)
        print('Unique stocks traded', str(len(self.perf.exchange_ticker.unique())))
        print('Stocks traded', self.perf.exchange_ticker.unique())
        b = self.perf
        b['date'] = pd.to_datetime(b['market_trade_date'], format='%Y-%m-%d')
        b = b.set_index(pd.DatetimeIndex(b['date']))
        # s = b.groupby(by=['trade_date','stock_id']).count() # this gives me a count for each stock each day

        b['stock_id'].resample("M").count()
        b['year'] = b.date.dt.year
        c = b.drop_duplicates(subset=['year', 'exchange_ticker'])

        u = c.groupby(['year'])['stock_id'].count()

        print(u)

        # calc the forward return
        b['value'] = pd.to_numeric(b.market_close_price) * pd.to_numeric(b.shares)
        # b['forward_month_ret'] = b.groupby(level=0)['tret'].shift(1)
        d = b
        d['value_prev'] = d.groupby(level=0)['value'].shift(1)
        d['return'] = (d.value / d.groupby(level=0)['value'].shift(1)) - 1

        for year in b['year'].unique():
            # print(year)
            stocks = b.query('(year == ' + str(year) + ')')['exchange_ticker'].unique()
            print(year, stocks)

        wb.save('/Users/connahcutbush/Desktop/perf_summary.xlsx')
